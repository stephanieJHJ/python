# *encoding=utf-8
from datetime import datetime
from openpyxl import Workbook
import time

CN_NUM = {
    1: '一', 2: '二', 3: '三', 4: '四', 5: '五', 6: '六', 7: '七', 8: '八', 9: '九', 0: ''
}
CN_UNIT = {
    10: '十',
    100: '百',
    1000: '千',
    10000: '万'
}

list1 = []
list2 = []
list3 = []

#转换大小写
def arabic_to_chinese(ara: int) -> str:
    global val
    if ara < 10:
        val = CN_NUM[ara]
    if 10 <= ara < 100:
        hv = int(ara / 10)
        hv2 = CN_NUM[hv]
        lv = ara % 10
        lv2 = CN_NUM[lv]
        if hv == 1:
            val = '十' + lv2
        else:
            val = hv2 + '十' + lv2
    if 100 <= ara < 1000:
        pass
    return val  #zhu  #

#输入不同时区，获取开始时间
def zone_start_time_choice(zone):
    tstart1 = 0
    if zone == "shenyang":
        tstart1 = 946667640
    if zone == "harbin":
        tstart1 = 946668060
    if zone == 'shanghai' or zone == 'dalian':
        tstart1 = 946667160
    return tstart1

#结果写入excel
def write_into_excel(list2, list1):
    wb = Workbook()  # 创建文件对象
    ws = wb.active  # 获取第一个sheet
    for i in range(len(list2)):
        ws.cell(i + 1, 1, list2[i])  # 写入水下N刻
    for i in range(len(list1)):
        ws.cell(i + 1, 2, list1[i])  # 写入时间
    # wb.save("{ }.xlsx".format(int(time.time)))
    wb.save("test.xlsx")


if __name__ == "__main__":

    tstart = zone_start_time_choice(input())

    for i in range(1, 101):
        m1 = datetime.fromtimestamp(tstart)  # 格式化初始时间
        cm1 = m1.strftime("%H:%M")
        tend = tstart + 864
        m2 = datetime.fromtimestamp(tend)  # 格式化后者时间
        cm2 = m2.strftime("%H:%M")
        result = cm1 + "————————" + cm2
        list1.append(result)  # 时间写入list1
        tstart = tend

        list3.append(i)  # 为转换大写创建数组
        for ara in list3:
            x = arabic_to_chinese(ara)  # 转化为大写
        info = "水下%s刻" % (x)
        list2.append(info)  # 水下N刻写入list2

    write_into_excel(list2,list1)


